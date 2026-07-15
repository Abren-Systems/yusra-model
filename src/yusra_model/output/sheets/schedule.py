"""Sheet 2: Loan Repayment Schedule"""
from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from yusra_model.models.loans import Portfolio

HDR_FILL = PatternFill("solid", fgColor="4472C4")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
TITLE_FONT = Font(bold=True, size=14, color="003366", name="Calibri")
SECTION_FONT = Font(bold=True, size=12, color="003366", name="Calibri")
VAL_FONT = Font(size=10, name="Calibri")
BOLD_VAL = Font(bold=True, size=10, name="Calibri")
FMT_NUM = '#,##0'
FMT_NUM2 = '#,##0.00'
FMT_DATE = 'YYYY-MM-DD'
FILL_BLUE = PatternFill("solid", fgColor="D6E4F0")
THIN = Border(left=Side('thin', 'C0C0C0'), right=Side('thin', 'C0C0C0'),
              top=Side('thin', 'C0C0C0'), bottom=Side('thin', 'C0C0C0'))


def hdr(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN


def write_row(ws, row, col_start, values, bold=False, fmt=None, fill=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=v)
        cell.font = BOLD_VAL if bold else VAL_FONT
        if fmt:
            if isinstance(fmt, list):
                if i < len(fmt) and fmt[i]:
                    cell.number_format = fmt[i]
            else:
                cell.number_format = fmt
        if fill:
            cell.fill = fill
        cell.border = THIN


def borders(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = THIN


def build(ws: openpyxl.worksheet.worksheet.Worksheet, cfg, portfolio: Portfolio) -> None:
    ws.title = "Loan_Repayment_Schedule"
    ws.sheet_properties.tabColor = "4472C4"
    for col, w in [(1, 3), (2, 24), (3, 14), (4, 16), (5, 16), (6, 16), (7, 12), (8, 18), (9, 18), (10, 18), (11, 18)]:
        ws.column_dimensions[chr(64 + col)].width = w

    ws.merge_cells('B1:K1')
    ws['B1'] = f"{cfg.company} — MURABAHA REVOLVING LOAN REPAYMENT SCHEDULE"
    ws['B1'].font = TITLE_FONT

    row_h = 3
    cols = ["Supplier", "USD Value", "ETB Principal", "Start Date", "End Date",
            "Quarter", "Quarterly Repayment\n(Total)", "Profit Charge", "Principal Repayment", "Balance\n(Total Deferred)"]
    hdr(ws, row_h, 2, 11)
    write_row(ws, row_h, 2, cols, bold=True)

    qlabels = [f"Q{q} {yr}" for yr in range(2026, 2029) for q in range(1, 5)]
    qlabels = qlabels[qlabels.index("Q2 2026"):qlabels.index("Q4 2028") + 1]

    curr_row = 4
    for loan in portfolio.loans:
        entries = loan.repayment_schedule(qlabels)
        for i, entry in enumerate(entries):
            vals = [entry["supplier"], entry["usd_value"], entry["etb_principal"],
                    entry["start_date"], entry["end_date"], entry["quarter"],
                    entry["quarterly_repayment"], entry["profit_charge"],
                    entry["principal_repayment"], entry["balance"]]
            fmts = [None, FMT_NUM, FMT_NUM, FMT_DATE, FMT_DATE, None, FMT_NUM, FMT_NUM2, FMT_NUM2, FMT_NUM2]
            write_row(ws, curr_row, 2, vals, fmt=fmts)
            if i == 0:
                for c in range(2, 12):
                    ws.cell(curr_row, c).fill = FILL_BLUE
            curr_row += 1
        curr_row += 1

    curr_row += 1
    ws.cell(curr_row, 2, "SCHEDULE SUMMARY").font = SECTION_FONT
    curr_row += 1

    rows = [
        ("Total Loan Principal Financed", portfolio.total_principal, FMT_NUM),
        ("Total Quarterly Repayment (all loans)", portfolio.total_quarterly_repayment, FMT_NUM),
        ("Total Profit Charges (8 quarters)", portfolio.total_profit, FMT_NUM2),
        ("Total Cost of Financing", portfolio.total_cost, FMT_NUM2),
    ]
    for lbl, val, nf in rows:
        write_row(ws, curr_row, 2, [lbl, val], bold=True, fmt=[None, nf])
        borders(ws, curr_row, curr_row, 2, 3)
        curr_row += 1
