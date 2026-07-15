"""Sheet 4: Liquidity Dashboard"""
from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from yusra_model.models.cashflow import CashFlowProjection, QUARTER_LABELS

ACCENT_FILL = PatternFill("solid", fgColor="BF8F00")
ACCENT_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
SECTION_FONT = Font(bold=True, size=12, color="BF8F00", name="Calibri")
TITLE_FONT = Font(bold=True, size=16, color="BF8F00", name="Calibri")
BOLD_VAL = Font(bold=True, size=10, name="Calibri")
VAL_FONT = Font(size=10, name="Calibri")
FMT_NUM = '#,##0'
FMT_NUM2 = '#,##0.00'
FMT_PCT = '0.0%'
FILL_GREEN = PatternFill("solid", fgColor="E2EFDA")
FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")
FILL_RED = PatternFill("solid", fgColor="FCE4EC")
FILL_KPI = PatternFill("solid", fgColor="FFF8E1")
THIN = Border(left=Side('thin', 'C0C0C0'), right=Side('thin', 'C0C0C0'),
              top=Side('thin', 'C0C0C0'), bottom=Side('thin', 'C0C0C0'))


def build(ws: openpyxl.worksheet.worksheet.Worksheet, cfg, portfolio, proj: CashFlowProjection) -> None:
    ws.title = "Liquidity_Dashboard"
    ws.sheet_properties.tabColor = "BF8F00"
    ws.sheet_view.showGridLines = False

    for col, w in [(1, 2), (2, 30), (3, 22), (4, 22), (5, 22), (6, 10), (7, 30), (8, 22), (9, 22), (10, 22)]:
        ws.column_dimensions[chr(64 + col)].width = w

    ws.merge_cells('B1:J1')
    ws['B1'] = "LIQUIDITY DASHBOARD — STRATEGIC PLANNING"
    ws['B1'].font = TITLE_FONT
    ws.merge_cells('B2:J2')
    ws['B2'] = f"{cfg.company} — Murabaha Revolving Loan Facility (ETB {cfg.total_facility:,.0f})"
    ws['B2'].font = Font(italic=True, size=11, color="666666", name="Calibri")

    # KPI row
    qlabels = QUARTER_LABELS
    first_row = proj.rows[0] if proj.rows else None

    kpi_start = 4
    kpis = [
        ("Net Cash Position\n(Current Quarter)", first_row.closing_cash if first_row else 0, FMT_NUM2, "Closing cash balance\n= Opening + Net CF"),
        ("Facility Utilization\n(Current)", first_row.facility_util_pct if first_row else 0, FMT_PCT, "Outstanding / 70M"),
        ("Total Drawdowns\n(All Loans)", portfolio.total_principal, FMT_NUM, "Total ETB principal drawn"),
        ("Inventory\nTurnover (Annualized)", "4x (3m) / 2x (6m)", None, "Depends on sales cycle"),
        ("Total Quarterly\nRepayment", portfolio.total_quarterly_repayment, FMT_NUM, "Sum of all 4 loan payments"),
    ]
    for ki, (label, val, nf, desc) in enumerate(kpis):
        kc = 2 + ki
        ws.cell(kpi_start - 1, kc, label).font = Font(bold=True, size=9, color="333333", name="Calibri")
        ws.cell(kpi_start - 1, kc).alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        cell = ws.cell(kpi_start, kc, val)
        if nf:
            cell.number_format = nf
        cell.font = Font(bold=True, size=16, color="BF8F00", name="Calibri")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = FILL_KPI
        cell.border = Border(left=Side('medium', 'BF8F00'), right=Side('medium', 'BF8F00'),
                             top=Side('medium', 'BF8F00'), bottom=Side('medium', 'BF8F00'))
        ws.cell(kpi_start + 1, kc, desc).font = Font(size=8, italic=True, color="888888", name="Calibri")
        ws.cell(kpi_start + 1, kc).alignment = Alignment(horizontal='center', wrap_text=True)

    # Quarterly summary table
    sum_row = kpi_start + 4
    ws.merge_cells(f'B{sum_row}:J{sum_row}')
    ws.cell(sum_row, 2, "QUARTERLY LIQUIDITY SUMMARY").font = SECTION_FONT
    sum_row += 1
    sum_headers = ["Quarter", "Closing Cash", "Net CF", "Sales Inflow", "Repayments", "Facility Util %", "Inventory"]
    for i, h in enumerate(sum_headers):
        cell = ws.cell(sum_row, 2 + i, h)
        cell.font = ACCENT_FONT; cell.fill = PatternFill("solid", fgColor="BF8F00")
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = THIN

    for qi, q in enumerate(qlabels):
        sr = sum_row + 1 + qi
        row = proj.rows[qi] if qi < len(proj.rows) else None
        if row is None:
            continue
        ws.cell(sr, 2, q).font = VAL_FONT; ws.cell(sr, 2).alignment = Alignment(horizontal='center'); ws.cell(sr, 2).border = THIN
        ws.cell(sr, 3, row.closing_cash).number_format = FMT_NUM2; ws.cell(sr, 3).border = THIN
        ws.cell(sr, 4, row.net_cash_flow).number_format = FMT_NUM2; ws.cell(sr, 4).border = THIN
        ws.cell(sr, 5, row.sales_inflow_3m).number_format = FMT_NUM; ws.cell(sr, 5).border = THIN
        ws.cell(sr, 6, row.repayments).number_format = FMT_NUM; ws.cell(sr, 6).border = THIN
        ws.cell(sr, 7, row.facility_util_pct).number_format = FMT_PCT; ws.cell(sr, 7).border = THIN
        ws.cell(sr, 8, row.closing_inventory).number_format = FMT_NUM; ws.cell(sr, 8).border = THIN

    # Charts
    chart_start = sum_row + 1 + len(qlabels) + 3

    # Chart 1: Cash Balance Trend
    chart1 = LineChart()
    chart1.title = "Cash Balance Trend (ETB)"
    chart1.style = 10; chart1.y_axis.title = "ETB"; chart1.x_axis.title = "Quarter"
    chart1.width = 20; chart1.height = 12
    data_ref = Reference(ws, min_col=3, min_row=sum_row, max_row=sum_row + len(qlabels))
    cats_ref = Reference(ws, min_col=2, min_row=sum_row + 1, max_row=sum_row + len(qlabels))
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.series[0].graphicalProperties.line.width = 25000
    chart1.legend = None
    ws.add_chart(chart1, f"B{chart_start}")

    # Chart 2: Facility Utilization
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Facility Utilization %"; chart2.y_axis.title = "Utilization %"
    chart2.y_axis.numFmt = '0%'
    chart2.width = 20; chart2.height = 12
    data_ref2 = Reference(ws, min_col=7, min_row=sum_row, max_row=sum_row + len(qlabels))
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref)
    chart2.legend = None
    ws.add_chart(chart2, f"B{chart_start + 16}")

    # Chart 3: Inventory vs Repayments Stacked
    chart3 = BarChart()
    chart3.type = "col"; chart3.grouping = "stacked"
    chart3.title = "Inventory Balance vs Loan Repayments (ETB)"; chart3.y_axis.title = "ETB"
    chart3.width = 20; chart3.height = 12
    data_ref3a = Reference(ws, min_col=8, min_row=sum_row, max_row=sum_row + len(qlabels))
    data_ref3b = Reference(ws, min_col=6, min_row=sum_row, max_row=sum_row + len(qlabels))
    chart3.add_data(data_ref3a, titles_from_data=True)
    chart3.add_data(data_ref3b, titles_from_data=True)
    chart3.set_categories(cats_ref)
    chart3.series[0].graphicalProperties.solidFill = "548235"
    chart3.series[1].graphicalProperties.solidFill = "4472C4"
    ws.add_chart(chart3, f"B{chart_start + 32}")

    # Strategic KPIs section
    skpi_row = chart_start + 48
    ws.cell(skpi_row, 2, "STRATEGIC KPIs — CAPITAL EFFICIENCY & LIQUIDITY").font = SECTION_FONT

    monthly_oh = cfg.overheads_per_month
    total_loan = cfg.total_facility
    loan_profit_monthly = total_loan * 2 * cfg.profit_rate / 24
    min_cash_buffer = monthly_oh * 3

    strategic_kpis = [
        ("Minimum Sales\nTarget (Monthly)", round((monthly_oh + loan_profit_monthly) / cfg.gross_profit_margin, 2), FMT_NUM, "Breakeven at 30% GM"),
        ("Cash Buffer\nRequirement", min_cash_buffer, FMT_NUM2, "3 months overheads"),
        ("Loan Profit\nper Month", round(loan_profit_monthly, 2), FMT_NUM2, f"{cfg.profit_rate*100:.0f}% x {total_loan:,.0f} x 2yr / 24mo"),
        ("Total Principal\nFinanced", portfolio.total_principal, FMT_NUM, "All 4 loans"),
        ("Total Profit\nCharges (8 qtrs)", portfolio.total_profit, FMT_NUM2, f"{cfg.profit_rate*100:.0f}% flat Murabaha"),
    ]
    for ki, (label, val, nf, desc) in enumerate(strategic_kpis):
        kc = 2 + ki
        ws.cell(skpi_row + 1, kc, label).font = Font(bold=True, size=9, color="002060")
        ws.cell(skpi_row + 1, kc).alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        cell = ws.cell(skpi_row + 2, kc, val)
        cell.number_format = nf
        cell.font = Font(bold=True, size=14, color="002060", name="Calibri")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill("solid", fgColor="E8EAF6")
        cell.border = Border(left=Side('medium', '002060'), right=Side('medium', '002060'),
                             top=Side('medium', '002060'), bottom=Side('medium', '002060'))
        ws.cell(skpi_row + 3, kc, desc).font = Font(size=8, italic=True, color="888888", name="Calibri")
        ws.cell(skpi_row + 3, kc).alignment = Alignment(horizontal='center', wrap_text=True)

    # Recycling KPIs
    recycle_row = skpi_row + 5
    ws.cell(recycle_row, 2, "RECYCLING KPIs").font = SECTION_FONT
    dr = recycle_row + 1
    recycling_kpis = [
        ("Total Initial\nThroughput", portfolio.total_principal, FMT_NUM, f"{len(portfolio.loans)} LCs at facility start"),
        ("CEO Throughput\nTarget", cfg.ceo_throughput_target, FMT_NUM, f"{cfg.ceo_throughput_target/cfg.total_facility:.1f}x facility utilisation"),
        ("Required Avg\nLC Tenor", "~2.3 qtrs", None, f"{cfg.ceo_throughput_target/cfg.total_facility:.1f} turns in 8 quarters"),
        ("Recycle\nEfficiency", f"{cfg.ceo_throughput_target/cfg.total_facility:.1f}x", None, f"{cfg.ceo_throughput_target/cfg.total_facility:.1f} = target ratio"),
        ("Required Sales\nCycle", "3 Months", None, "Critical for velocity"),
    ]
    for i, (lbl, val, nf, desc) in enumerate(recycling_kpis):
        kc = 2 + i
        is_str = "CEO" in lbl or "240" in str(val)
        ws.cell(dr - 1, kc, lbl).font = Font(bold=True, size=9, color="002060", name="Calibri")
        ws.cell(dr - 1, kc).alignment = Alignment(horizontal='center', wrap_text=True, vertical='bottom')
        cell = ws.cell(dr, kc, val)
        if nf:
            cell.number_format = nf
        cell.fill = PatternFill("solid", fgColor="FCE4EC" if is_str else "E8EAF6")
        cell.font = Font(bold=True, size=14, color="C00000" if is_str else "002060", name="Calibri")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side('medium', '002060'), right=Side('medium', '002060'),
                             top=Side('medium', '002060'), bottom=Side('medium', '002060'))
        ws.cell(dr + 1, kc, desc).font = Font(size=8, italic=True, color="888888", name="Calibri")
        ws.cell(dr + 1, kc).alignment = Alignment(horizontal='center', wrap_text=True)
