"""Sheet 5: Strategic Plan & Targets — driven by optimizer, not CEO desires."""
from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from yusra_model.models.targets import build_targets, build_velocity_scenarios
from yusra_model.models.optimizer import optimize

FONT_ACCENT  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
FONT_SECTION = Font(bold=True, size=12, color="002060", name="Calibri")
FONT_BOLD    = Font(bold=True, size=10, name="Calibri")
FONT_14R     = Font(bold=True, size=14, color="C00000", name="Calibri")
FONT_12B     = Font(bold=True, size=12, color="002060", name="Calibri")
FONT_10_ITAL = Font(italic=True, size=10, color="555555", name="Calibri")
FONT_10_GREY = Font(italic=True, size=10, color="888888", name="Calibri")
FONT_BOLD_C  = Font(bold=True, size=10, color="C00000", name="Calibri")
FONT_12B_C   = Font(bold=True, size=12, color="C00000", name="Calibri")
FONT_14B_C   = Font(bold=True, size=14, color="C00000", name="Calibri")
FONT_9       = Font(size=9, name="Calibri")
FONT_9_ITAL  = Font(size=9, italic=True, name="Calibri")

GP   = PatternFill("solid", fgColor="E2EFDA")
YP   = PatternFill("solid", fgColor="FFF2CC")
BP   = PatternFill("solid", fgColor="D6E4F0")
RP   = PatternFill("solid", fgColor="FCE4EC")
GOLP = PatternFill("solid", fgColor="FFF8E1")
HP   = PatternFill("solid", fgColor="002060")

THIN  = Border(left=Side('thin','C0C0C0'), right=Side('thin','C0C0C0'),
               top=Side('thin','C0C0C0'), bottom=Side('thin','C0C0C0'))
AC    = Alignment(horizontal='center', wrap_text=True)
ACT   = Alignment(horizontal='center')
AWT   = Alignment(wrap_text=True)


def build(ws, cfg, portfolio):
    ws.title = "Strategic_Plan"
    ws.sheet_properties.tabColor = "002060"
    ws.sheet_view.showGridLines = False
    for col, w in [(1, 2), (2, 30), (3, 24), (4, 50), (5, 22), (6, 22)]:
        ws.column_dimensions[chr(64 + col)].width = w

    ws.merge_cells('B1:F1')
    ws['B1'].value = "STRATEGIC PLAN - OPTIMAL VS ASPIRATIONAL TARGETS"
    ws['B1'].font = Font(bold=True, size=14, color="002060", name="Calibri")

    ws.merge_cells('B2:F2')
    ws['B2'].value = f"{cfg.company} | Data-Driven Target Optimisation"
    ws['B2'].font = Font(italic=True, size=10, color="666666")

    opt = optimize(cfg)
    targets = build_targets(cfg)
    scenarios = build_velocity_scenarios(cfg)

    r = 4
    _tr(ws, r, "PRIMARY OBJECTIVE: MAXIMISE LOAN RECYCLING WITHIN CONSTRAINTS", FONT_SECTION)
    r += 1
    binding = opt.optimum.binding_constraint
    _tr(ws, r,
        f"Computed optimal throughput: {opt.optimum.max_throughput:,.0f} ETB ({opt.optimum.max_multiplier:.1f}x facility) "
        f"at {opt.optimum.optimal_tenor}-quarter tenors. Bound by '{binding}'. "
        f"Aspirational target {cfg.ceo_throughput_target:,.0f} exceeds feasible maximum by {opt.gap_to_aspirational:,.0f}.",
        FONT_10_ITAL)
    r += 2

    _tr(ws, r, f"OPTIMAL THROUGHPUT: ETB {opt.optimum.max_throughput:,.0f} ({opt.optimum.max_multiplier:.1f}x FACILITY)", FONT_14B_C, RP, AC)
    r += 1
    _tr(ws, r, f"Tenor: {opt.optimum.optimal_tenor}-quarters | Constraint: {binding} | Breakeven: {opt.breakeven_throughput:,.0f}", FONT_10_GREY)
    r += 1

    for i, h in enumerate(['Step', 'Amount', 'How', 'Cumulative']):
        _hdr(ws, r, 2 + i, h)
    r += 1

    feasible = [s for s in opt.solutions if s.feasible]
    steps_data = []
    if feasible:
        best = feasible[-1]
        best_t = best.tenor_quarters
        for si, sol in enumerate(feasible):
            label = f'{si+1}. {sol.tenor_quarters}-quarter cycle'
            amt = f'ETB {sol.throughput:,.0f}'
            desc = f'{sol.multiplier:.1f}x turnover, min closing cash ETB {sol.min_closing_cash:,.0f}'
            cum = int(sol.throughput)
            pf = GP if si == len(feasible) - 1 else BP
            steps_data.append((label, amt, desc, cum, pf))
    else:
        steps_data.append(('All tenors infeasible', 'ETB 0', 'No feasible recycling path within constraints', 0, RP))

    for lbl, amt, desc, cum, pf in steps_data:
        _cv(ws, r, 2, lbl, FONT_BOLD, pf, THIN)
        _cv(ws, r, 3, amt, Font(bold=True, size=12 if pf == GP else 10, color='C00000' if pf == GP else '002060', name='Calibri'), pf, THIN, ACT)
        _cv(ws, r, 4, desc, FONT_9, pf, THIN, AWT)
        _cv(ws, r, 5, '', None, pf, THIN)
        _cv(ws, r, 6, cum, FONT_12B, pf, THIN, ACT, '#,##0')
        r += 1

    r += 1

    _tr(ws, r, f"CONSTRAINT BREAKDOWN", FONT_SECTION)
    r += 1
    constraints = _build_constraint_rows(opt)
    for i, h in enumerate(['Constraint', 'Status', 'Detail', 'Recommendation']):
        _hdr(ws, r, 2 + i, h)
    r += 1
    for con in constraints:
        pf = RP if 'breach' in con['status'].lower() or 'binding' in con['status'].lower() else GP
        _cv(ws, r, 2, con['name'], FONT_BOLD, pf, THIN)
        _cv(ws, r, 3, con['status'], Font(bold=True, color="C00000" if pf == RP else "002060"), pf, THIN, ACT)
        _cv(ws, r, 4, con['detail'], FONT_9, pf, THIN, AWT)
        _cv(ws, r, 5, con['recommendation'], FONT_9_ITAL, pf, THIN, AWT)
        r += 1

    r += 1
    _tr(ws, r, "VELOCITY SCENARIO COMPARISON (OPTIMISER SWEEP)", FONT_SECTION)
    r += 1
    for i, h in enumerate(['Scenario', 'Avg LC Tenor', 'Turns in 8 Qtrs', 'Throughput', 'Feasibility']):
        _hdr(ws, r, 2 + i, h)
    r += 1

    for sc in scenarios:
        is_asp = 'Aspirational' in sc.scenario
        pf = RP if 'not achievable' in sc.feasibility or 'removing' in sc.feasibility else (GP if is_asp else BP)
        if not is_asp and 'Constrained' in sc.feasibility:
            pf = YP
        f_name = FONT_BOLD_C if is_asp else FONT_BOLD
        f_thr = Font(bold=True, size=12 if is_asp else 10, color='C00000' if is_asp else '002060', name='Calibri')
        _cv(ws, r, 2, sc.scenario, f_name, pf, THIN)
        _cv(ws, r, 3, sc.avg_tenor, FONT_BOLD, pf, THIN, ACT)
        _cv(ws, r, 4, sc.turns, FONT_BOLD, pf, THIN, ACT)
        _cv(ws, r, 5, sc.throughput, f_thr, pf, THIN, ACT)
        _cv(ws, r, 6, sc.feasibility, FONT_9_ITAL, pf, THIN, AC)
        r += 1

    r += 1
    _tr(ws, r, "RECYCLING TARGETS — OPTIMISER OUTPUT", FONT_SECTION)
    r += 1
    for lbl, val, mult, method, pf in [
        ('Breakeven (Minimum Viable)', f'ETB {opt.breakeven_throughput:,.0f}', f'{opt.breakeven_multiplier:.1f}x', 'Covers overheads + loan profit charges', YP),
        ('Optimal (Constraint-Bounded)', f'ETB {opt.optimum.max_throughput:,.0f}', f'{opt.optimum.max_multiplier:.1f}x', f'{opt.optimum.optimal_tenor}-quarter tenors, bound by {binding}', GP),
        ('Passive (No Recycling)', f'ETB {cfg.total_facility:,.0f}', '1.0x', 'Single pass, 8-quarter tenors, no redraws', BP),
    ]:
        _cv(ws, r, 2, lbl, FONT_BOLD, pf, THIN)
        _cv(ws, r, 3, val, Font(bold=True, size=14 if lbl.startswith('Optimal') else 12, color='C00000' if lbl.startswith('Optimal') else '002060', name='Calibri'), pf, THIN, ACT)
        _cv(ws, r, 4, mult, FONT_BOLD, pf, THIN, ACT)
        _cv(ws, r, 5, method, FONT_9, pf, THIN, AWT)
        _cv(ws, r, 6, '' if lbl.startswith('Optimal') else '', None, pf, THIN)
        r += 1

    r += 2
    _tr(ws, r, "OPERATIONAL PHASES — DRIVEN BY OPTIMAL THROUGHPUT", FONT_SECTION)
    r += 1
    best_tenor = opt.optimum.optimal_tenor
    if feasible:
        best_sol = feasible[-1]
        phases = [
            ('Phase 1: Initial deployment', f'ETB {cfg.total_facility:,.0f} drawn',
             f'Full facility deployed as {best_tenor}-quarter Murabaha LCs. Inventory purchased.'),
            ('Phase 2: Redraw repaid principal', f'~ETB {best_sol.throughput - cfg.total_facility:,.0f} recycled',
             f'Principal repaid frees capacity. Redraw immediately as {best_tenor}-quarter LCs. Repeat each cycle.'),
            ('Phase 3: Full velocity achieved', f'ETB {best_sol.throughput:,.0f} total throughput',
             f'{best_sol.multiplier:.1f}x facility utilisation. Min closing cash ETB {best_sol.min_closing_cash:,.0f} — within constraints.'),
        ]
    else:
        phases = [('No feasible path', '—', 'Relax constraints or increase facility/funding.')]
    for ph, amt, desc in phases:
        ws.cell(r, 2, ph).font = Font(bold=True, size=10, color="002060", name="Calibri")
        ws.cell(r, 3, amt).font = FONT_BOLD
        ws.cell(r, 3).alignment = ACT
        ws.cell(r, 4, desc).font = FONT_9
        ws.cell(r, 4).alignment = AWT
        r += 1


def _build_constraint_rows(opt) -> list[dict]:
    binding = opt.optimum.binding_constraint
    rows = [
        dict(name='DSCR (≥1.2)', status='Pass' if opt.optimum.dscr_ok else 'Binding',
             detail=f'DSCR={opt.optimum.dscr_at_limit:.2f}' if not opt.optimum.dscr_ok else 'No constraint breach',
             recommendation='Extend loan tenor to reduce quarterly repayments' if not opt.optimum.dscr_ok else 'Maintain current tenor'),
        dict(name='Cash Buffer (min ETB 0)', status='Pass' if opt.optimum.min_cash >= 0 else 'Binding',
             detail=f'Min closing cash ETB {opt.optimum.min_cash:,.0f}',
             recommendation='Increase equity contribution or reduce facility draw' if opt.optimum.min_cash < 0 else 'OK'),
        dict(name='Facility Ceiling', status='Pass' if opt.optimum.max_principal <= opt.optimum.facility else 'Binding',
             detail=f'Max principal ETB {opt.optimum.max_principal:,.0f} vs ceiling ETB {opt.optimum.facility:,.0f}',
             recommendation='Request higher facility limit or reduce throughput velocity' if opt.optimum.max_principal > opt.optimum.facility else 'OK'),
    ]
    for r in rows:
        if r['name'].startswith(binding):
            r['status'] = 'Binding'
    return rows


def _tr(ws, r, text, font, fill=None, align=None):
    for c in range(2, 7):
        cell = ws.cell(r, c)
        cell.value = text if c == 2 else ''
        cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align


def _hdr(ws, r, c, text):
    cell = ws.cell(r, c, text)
    cell.font = FONT_ACCENT
    cell.fill = HP
    cell.alignment = AC
    cell.border = THIN


def _cv(ws, r, c, value, font, fill, border, align=None, nf=None):
    cell = ws.cell(r, c, value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if align:
        cell.alignment = align
    if nf:
        cell.number_format = nf
