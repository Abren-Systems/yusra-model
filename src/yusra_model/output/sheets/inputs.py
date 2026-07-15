"""Sheet 1: Inputs"""
from __future__ import annotations
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from typing import Optional
from yusra_model.config.loader import Config


def _murabaha_apr_equivalent(flat_rate: float, tenor_quarters: int) -> float:
    """Compute declining-balance APR equivalent for a flat Murabaha rate.

    Solves for the periodic rate r such that equal installments match.
    Flat total profit = principal * rate * years.
    Equivalent APR is found via bisection on the amortisation formula.
    """
    n = tenor_quarters
    # Flat-rate total profit factor per unit principal
    flat_factor = 1.0 + flat_rate * (n / 4)
    quarterly_pmt = flat_factor / n

    # Bisection to find quarterly rate that produces same payment
    lo, hi = 0.0, 0.5
    for _ in range(50):
        mid = (lo + hi) / 2
        f = quarterly_pmt - (mid * (1 + mid) ** n) / ((1 + mid) ** n - 1) if mid > 1e-12 else quarterly_pmt - 1.0 / n
        if f > 0:
            lo = mid
        else:
            hi = mid
    quarterly_rate = (lo + hi) / 2
    apr_eq = quarterly_rate * 4
    return round(apr_eq * 100, 1)

HDR_FILL = PatternFill("solid", fgColor="003366")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
ACCENT_FILL = PatternFill("solid", fgColor="4472C4")
LABEL_FONT = Font(bold=True, size=10, name="Calibri")
VAL_FONT = Font(size=10, name="Calibri")
BOLD_VAL = Font(bold=True, size=10, name="Calibri")
TITLE_FONT = Font(bold=True, size=14, color="003366", name="Calibri")
SECTION_FONT = Font(bold=True, size=12, color="003366", name="Calibri")
FMT_NUM = '#,##0'
FMT_NUM2 = '#,##0.00'
FMT_PCT = '0.0%'
FMT_DATE = 'YYYY-MM-DD'
FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(left=Side('thin', 'C0C0C0'), right=Side('thin', 'C0C0C0'),
              top=Side('thin', 'C0C0C0'), bottom=Side('thin', 'C0C0C0'))


def hdr(ws, row, start_col, end_col, fill=HDR_FILL, font=HDR_FONT):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN


def borders(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = THIN


def write_row(ws, row, col_start, values, bold=False, fmt=None, fill=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=v)
        cell.font = BOLD_VAL if bold else VAL_FONT
        if fmt:
            if isinstance(fmt, list):
                if i < len(fmt) and fmt[i]:
                    cell.number_format = fmt[i]
            else:
                cell.number_format = fmt
        if fill:
            cell.fill = fill
        cell.border = THIN


def build(ws: openpyxl.worksheet.worksheet.Worksheet, cfg: Config, portfolio,
          audit: dict | None = None) -> None:
    ws.title = "Inputs"
    ws.sheet_properties.tabColor = "003366"
    for col, w in [(1, 3), (2, 42), (3, 22), (4, 22), (5, 22), (6, 22), (7, 18), (8, 22), (9, 22)]:
        ws.column_dimensions[chr(64 + col)].width = w

    r = 1
    ws.merge_cells('B1:H1')
    ws['B1'] = f"{cfg.company} — FINANCIAL MODEL INPUT PARAMETERS"
    ws['B1'].font = TITLE_FONT

    # Audit info row
    if audit:
        ws.cell(r + 1, 2,
                f"Run: {audit.get('run_id', '?')} | "
                f"{audit.get('timestamp', '?')[:19]} | "
                f"Scenario: {audit.get('scenario', '?')} | "
                f"Model v{audit.get('model_version', '?')}").font = Font(size=9, italic=True, color="888888")

    r = 3
    ws.cell(r, 2, "KEY FINANCIAL PARAMETERS").font = SECTION_FONT
    params = [
        ("Cash Opening Balance (ETB)", cfg.opening_cash, FMT_NUM2),
        ("Inventory Opening Balance (ETB)", cfg.opening_inventory, FMT_NUM),
        ("Total Loan Facility (ETB)", cfg.total_facility, FMT_NUM),
        ("Overheads per Month (ETB)", cfg.overheads_per_month, FMT_NUM2),
        ("Annual Profit Charge Rate (Murabaha)", cfg.profit_rate, FMT_PCT),
        ("Loan Tenor (Quarters)", cfg.loan_tenor_quarters, '0'),
    ]
    for i, (lbl, val, nf) in enumerate(params):
        rr = r + i
        ws.cell(rr, 2, lbl).font = LABEL_FONT
        ws.cell(rr, 3, val).number_format = nf
        ws.cell(rr, 3).font = BOLD_VAL
        ws.cell(rr, 2).border = THIN
        ws.cell(rr, 3).border = THIN
    borders(ws, 4, 9, 2, 3)

    r = 9
    ws.cell(r, 2, "Overheads per Quarter (ETB)").font = LABEL_FONT
    ws.cell(r, 3, cfg.overheads_per_month * 3).number_format = FMT_NUM2
    ws.cell(r, 3).font = BOLD_VAL
    ws.cell(r, 2).border = THIN; ws.cell(r, 3).border = THIN

    r = 12
    ws.cell(r, 2, "EXCHANGE RATE SCENARIOS").font = SECTION_FONT
    r = 13
    hdr(ws, r, 2, 5, ACCENT_FILL)
    write_row(ws, r, 2, ["Scenario", "Rate (ETB/USD)", "Selected", "Description"], bold=True)
    r = 14
    write_row(ws, r, 2, ["Baseline", cfg.baseline_rate, "Yes", "Current market rate assumption"])
    ws.cell(r, 3).number_format = '#,##0'
    r = 15
    write_row(ws, r, 2, ["Stress", cfg.stress_rate, "", "Depreciation / devaluation scenario"])
    ws.cell(r, 3).number_format = '#,##0'

    r = 17
    ws.cell(r, 2, "Active Exchange Rate (ETB/USD)").font = LABEL_FONT
    ws.cell(r, 3, f"=IF(D14=\"Yes\",C14,C15)").number_format = '#,##0'
    ws.cell(r, 3).font = Font(bold=True, size=12, color="CC0000", name="Calibri")
    ws.cell(r, 2).border = THIN; ws.cell(r, 3).border = THIN

    r = 20
    ws.cell(r, 2, "SALES & PAYMENT ASSUMPTIONS").font = SECTION_FONT
    r = 21
    hdr(ws, r, 2, 5, ACCENT_FILL)
    write_row(ws, r, 2, ["Parameter", "Value", "Unit / Notes"], bold=True)
    r = 22
    write_row(ws, r, 2, ["Sales Cycle Duration", "3 Months", 'Enter "3 Months" or "6 Months"'])
    r = 23
    write_row(ws, r, 2, ["Cash Sales (Immediate)", cfg.cash_ratio, "% of sales collected immediately"])
    ws.cell(r, 3).number_format = FMT_PCT
    r = 24
    write_row(ws, r, 2, ["Credit Sales (30-day lag)", cfg.credit_ratio, "% collected following month/quarter"])
    ws.cell(r, 3).number_format = FMT_PCT
    r = 25
    write_row(ws, r, 2, ["Gross Profit Margin on Sales", cfg.gross_profit_margin, "Markup on cost of goods sold"])
    ws.cell(r, 3).number_format = FMT_PCT

    r = 28
    ws.cell(r, 2, "LOAN-FINANCED IMPORT PIPELINE").font = SECTION_FONT
    r = 29
    loan_headers = ["Supplier", "USD Value", "Effective Rate", "ETB Principal",
                    "Start Date", "Tenor (Qtrs)", "Quarterly Repayment", "Maturity Date"]
    hdr(ws, r, 2, 9, ACCENT_FILL)
    write_row(ws, r, 2, loan_headers, bold=True)

    for i, loan in enumerate(portfolio.loans):
        rr = r + 1 + i
        from datetime import timedelta
        mdt = date(loan.start_date.year + 2, loan.start_date.month, loan.start_date.day) if loan.start_date else ""
        vals = [loan.supplier, loan.usd_value, loan.effective_rate, loan.etb_principal,
                loan.start_date, loan.tenor_quarters, loan.quarterly_repayment, mdt]
        fmts = [None, FMT_NUM, '#,##0.00', FMT_NUM, FMT_DATE, '0', FMT_NUM, FMT_DATE]
        write_row(ws, rr, 2, vals, fmt=fmts)
        if i == 0:
            for c in range(3, 10):
                ws.cell(rr, c).fill = FILL_YELLOW

    r_tot = r + 1 + len(portfolio.loans)
    write_row(ws, r_tot, 2, ["TOTAL", "", "", round(portfolio.total_principal, 2), "", "", round(portfolio.total_quarterly_repayment, 2), ""], bold=True)
    ws.cell(r_tot, 4).number_format = FMT_NUM
    ws.cell(r_tot, 7).number_format = FMT_NUM
    borders(ws, r_tot, r_tot, 2, 9)
    borders(ws, 29, r_tot, 2, 9)

    r = r_tot + 2
    ws.cell(r, 2, "NOTES:").font = Font(bold=True, italic=True, size=10, color="666666")
    # APR-equivalent disclosure for Murabaha flat rate
    apr_eq = _murabaha_apr_equivalent(cfg.profit_rate, cfg.loan_tenor_quarters)
    notes = [
        "1. Reyoung LC already settled (Apr 7, 2026); stock in warehouse.",
        "2. Remaining LCs expected to settle in Oct 2026. Drawdowns modelled accordingly.",
        f"3. Overheads: ETB {cfg.overheads_per_month:,.2f}/month (approx {cfg.overheads_per_month*3:,.2f}/quarter).",
        "4. To switch scenarios: enter 'Yes' in D14 (Baseline) or D15 (Stress).",
        f"5. All profit charges calculated at {cfg.profit_rate*100:.0f}% p.a. using flat Murabaha structure.",
        f"   APR-equivalent (declining-balance): ~{apr_eq:.1f}%. This is structural, not additional cost.",
    ]
    for i, n in enumerate(notes):
        ws.cell(r + 1 + i, 2, n).font = Font(size=9, italic=True, color="888888")
