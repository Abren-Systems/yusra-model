"""Sheet 3: Quarterly Projection"""
from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from yusra_model.models.cashflow import CashFlowProjection, QUARTER_LABELS

HDR_FILL = PatternFill("solid", fgColor="548235")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
TITLE_FONT = Font(bold=True, size=14, color="003366", name="Calibri")
VAL_FONT = Font(size=10, name="Calibri")
BOLD_VAL = Font(bold=True, size=10, name="Calibri")
FMT_NUM = '#,##0'
FMT_NUM2 = '#,##0.00'
FMT_PCT = '0.0%'
FILL_GREEN = PatternFill("solid", fgColor="E2EFDA")
FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(left=Side('thin', 'C0C0C0'), right=Side('thin', 'C0C0C0'),
              top=Side('thin', 'C0C0C0'), bottom=Side('thin', 'C0C0C0'))


def hdr(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN


def write_row(ws, row, col_start, values, bold=False, fmt=None, fill=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=v)
        cell.font = BOLD_VAL if bold else VAL_FONT
        if fmt:
            if isinstance(fmt, list):
                if i < len(fmt) and fmt[i]:
                    cell.number_format = fmt[i]
            else:
                cell.number_format = fmt
        if fill:
            cell.fill = fill
        cell.border = THIN


def borders(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = THIN


def build(ws: openpyxl.worksheet.worksheet.Worksheet, cfg, portfolio, proj: CashFlowProjection) -> None:
    ws.title = "Quarterly_Projection"
    ws.sheet_properties.tabColor = "548235"
    col_widths = [3, 12, 18, 18, 18, 18, 16, 16, 18, 18, 20, 20]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.merge_cells('B1:L1')
    ws['B1'] = "YUSRA PHARMA PLC — QUARTERLY CASH FLOW PROJECTION"
    ws['B1'].font = TITLE_FONT

    ws.cell(2, 10, "Active Scenario:").font = Font(bold=True, italic=True, size=9)
    ws.cell(2, 11, "='Inputs'!C22").font = Font(bold=True, italic=True, size=9, color="4472C4")

    qcols = ["Quarter", "Opening Cash\nBalance", "Loan\nDrawdowns", "Loan\nRepayments",
             "Sales Inflows\n(3m Cycle)", "Sales Inflows\n(6m Cycle)", "Overheads",
             "Net Cash\nFlow", "Closing Cash\nBalance", "Closing Inventory\nBalance", "Facility Util.\n%"]

    row_h = 4
    hdr(ws, row_h, 2, 12)
    write_row(ws, row_h, 2, qcols, bold=True)

    qlabels = QUARTER_LABELS
    for qi, q in enumerate(qlabels):
        rr = 5 + qi
        row = proj.rows[qi] if qi < len(proj.rows) else None
        if row is None:
            continue
        ws.cell(rr, 2, q).font = BOLD_VAL
        ws.cell(rr, 2).alignment = Alignment(horizontal='center')
        ws.cell(rr, 3, row.opening_cash).number_format = FMT_NUM2; ws.cell(rr, 3).font = VAL_FONT
        ws.cell(rr, 4, row.drawdowns).number_format = FMT_NUM; ws.cell(rr, 4).font = VAL_FONT
        ws.cell(rr, 5, row.repayments).number_format = FMT_NUM2; ws.cell(rr, 5).font = VAL_FONT
        ws.cell(rr, 6, row.sales_inflow_3m).number_format = FMT_NUM; ws.cell(rr, 6).font = VAL_FONT
        ws.cell(rr, 7, row.sales_inflow_6m).number_format = FMT_NUM; ws.cell(rr, 7).font = VAL_FONT
        ws.cell(rr, 8, row.overheads).number_format = FMT_NUM2; ws.cell(rr, 8).font = VAL_FONT
        ws.cell(rr, 9, row.net_cash_flow).number_format = FMT_NUM2; ws.cell(rr, 9).font = VAL_FONT
        ws.cell(rr, 10, row.closing_cash).number_format = FMT_NUM2; ws.cell(rr, 10).font = BOLD_VAL
        ws.cell(rr, 11, row.closing_inventory).number_format = FMT_NUM; ws.cell(rr, 11).font = VAL_FONT
        ws.cell(rr, 12, row.facility_util_pct).number_format = FMT_PCT; ws.cell(rr, 12).font = BOLD_VAL

        # Color code scenario columns
        ws.cell(rr, 6).fill = FILL_GREEN
        ws.cell(rr, 7).fill = FILL_YELLOW

    borders(ws, 5, 5 + len(qlabels) - 1, 2, 12)

    nr = 5 + len(qlabels) + 1
    ws.cell(nr, 2, "SCENARIO SWITCH:").font = Font(bold=True, size=9, color="4472C4")
    ws.cell(nr, 3, 'Change Inputs!B22 between "3 Months" and "6 Months" to swap sales cycle.').font = Font(italic=True, size=9, color="888888")
    nr += 1
    ws.cell(nr, 2, "FORMULA NOTES:").font = Font(bold=True, size=9, color="4472C4")
    ws.cell(nr, 3, "Data computed from parameters. Adjust inputs in Inputs sheet to update.").font = Font(italic=True, size=9, color="888888")
