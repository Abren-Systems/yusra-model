"""Excel workbook orchestrator"""
from __future__ import annotations
from pathlib import Path
import openpyxl
from yusra_model.config.loader import Config
from yusra_model.models.loans import Loan, Portfolio
from yusra_model.models.cashflow import project
from yusra_model.output.sheets import inputs, schedule, projection, dashboard, strategy


def build_workbook(cfg: Config, output_path: str | Path) -> str:
    """Build the full 5-sheet Excel workbook."""
    # Build portfolio from config
    loans = [Loan(**l) for l in cfg.loans]
    portfolio = Portfolio(loans, cfg.total_facility, cfg.profit_rate)
    portfolio.allocate_remaining()

    # Run projection
    overhead_q = cfg.overheads_per_month * 3
    proj = project(portfolio, cfg.opening_cash, cfg.opening_inventory, overhead_q, cfg.gross_profit_margin)

    # Create workbook
    wb = openpyxl.Workbook()

    # Build sheets
    inputs.build(wb.active, cfg, portfolio)
    schedule.build(wb.create_sheet(), cfg, portfolio)
    projection.build(wb.create_sheet(), cfg, portfolio, proj)
    dashboard.build(wb.create_sheet(), cfg, portfolio, proj)
    strategy.build(wb.create_sheet(), cfg, portfolio)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return str(output_path)
