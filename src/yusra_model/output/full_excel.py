"""Excel workbook for full three-statement financial projection."""
from __future__ import annotations
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from yusra_model.engine import FinancialProjection

FMT_ACCENT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
FMT_TITLE  = Font(bold=True, size=14, color="002060", name="Calibri")
FMT_SECTION = Font(bold=True, size=12, color="002060", name="Calibri")
FMT_BOLD   = Font(bold=True, size=10, name="Calibri")
FMT_NUM    = '#,##0'
FILL_HDR   = PatternFill("solid", fgColor="002060")
FILL_ALT   = PatternFill("solid", fgColor="F2F7FB")
THIN = Border(left=Side('thin', 'C0C0C0'), right=Side('thin', 'C0C0C0'),
              top=Side('thin', 'C0C0C0'), bottom=Side('thin', 'C0C0C0'))
AC = Alignment(horizontal='center', wrap_text=True)


def build_full_workbook(
    cfg,
    proj: FinancialProjection,
    audit: dict | None = None,
    output_path: str | Path = "./reports/Full_Financial_Statements.xlsx",
) -> str:
    """Build an Excel workbook with full financial statements."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    _build_pnl(wb.active, proj, cfg)
    _build_balance_sheet(wb.create_sheet(), proj, cfg)
    _build_cash_flow(wb.create_sheet(), proj, cfg)
    _build_overview(wb.create_sheet(), proj, cfg, audit)

    wb.save(str(output_path))
    return str(output_path)


def _col(ws, widths: dict[int, int]):
    for c, w in widths.items():
        ws.column_dimensions[chr(64 + c)].width = w


def _hdr(ws, r, col_start, labels):
    for i, lbl in enumerate(labels):
        c = ws.cell(r, col_start + i, lbl)
        c.font = FMT_ACCENT
        c.fill = FILL_HDR
        c.alignment = AC
        c.border = THIN


def _row(ws, r, col_start, values, bold=False, nf=None, alt=False):
    for i, v in enumerate(values):
        c = ws.cell(r, col_start + i, v)
        c.font = FMT_BOLD if bold else Font(size=10, name="Calibri")
        if alt:
            c.fill = FILL_ALT
        c.border = THIN
        c.alignment = Alignment(horizontal='center' if i > 0 else 'left')
        if nf and i > 0:
            c.number_format = nf


def _build_pnl(ws, proj, cfg):
    ws.title = "Income_Statement"
    ws.sheet_properties.tabColor = "002060"
    _col(ws, {1: 3, 2: 30, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16})

    ws.merge_cells('B2:G2')
    ws['B2'] = f"{cfg.company} — Income Statement (5-Year Projection)"
    ws['B2'].font = FMT_TITLE

    heads = ["Metric"] + [str(y) for y in sorted(set(p.year for p in proj.periods))]
    r = 4
    _hdr(ws, r, 2, heads)

    ann = _annualize(proj, "income")
    lines = [
        ("Revenue", "revenue", True),
        ("Cost of Goods Sold", "cogs", False),
        ("Gross Profit", "gross_profit", True),
        ("", None, False),
        ("Operating Expenses:", None, True),
        ("  Sales & Marketing", "opex_sales_marketing", False),
        ("  Distribution", "opex_distribution", False),
        ("  Admin", "opex_admin", False),
        ("  R&D", "opex_r_and_d", False),
        ("  Other OpEx", "opex_other", False),
        ("Total Operating Expenses", "total_opex", True),
        ("EBITDA", "ebitda", True),
        ("Depreciation & Amortisation", "depreciation", False),
        ("EBIT", "ebit", True),
        ("Interest Expense", "interest_expense", False),
        ("EBT", "ebt", True),
        ("Tax Expense", "tax_expense", False),
        ("Net Income", "net_income", True),
    ]
    r += 1
    for i, (label, attr, is_bold) in enumerate(lines):
        if not attr:
            r += 1
            continue
        vals = [label] + [round(ann[y][attr], 0) for y in sorted(ann)]
        _row(ws, r, 2, vals, bold=is_bold, nf=FMT_NUM, alt=i % 2 == 0)
        r += 1


def _build_balance_sheet(ws, proj, cfg):
    ws.title = "Balance_Sheet"
    ws.sheet_properties.tabColor = "548235"
    _col(ws, {1: 3, 2: 30, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16})

    ws.merge_cells('B2:G2')
    ws['B2'] = f"{cfg.company} — Balance Sheet (5-Year Projection)"
    ws['B2'].font = FMT_TITLE

    heads = ["Metric"] + [str(y) for y in sorted(set(p.year for p in proj.periods))]
    r = 4
    _hdr(ws, r, 2, heads)

    # Use year-end (Q4) balances
    ye = _year_end(proj)

    lines = [
        ("ASSETS", None, True),
        ("Cash", "cash", False),
        ("Accounts Receivable", "accounts_receivable", False),
        ("Inventory", "inventory", False),
        ("Fixed Assets (Net)", "fixed_assets_net", False),
        ("Total Assets", "total_assets", True),
        ("", None, False),
        ("LIABILITIES", None, True),
        ("Accounts Payable", "accounts_payable", False),
        ("Long-Term Debt", "long_term_debt", False),
        ("Other Liabilities", "other_liabilities", False),
        ("Total Liabilities", "total_liabilities", True),
        ("", None, False),
        ("EQUITY", None, True),
        ("Paid-In Capital", "paid_in_capital", False),
        ("Retained Earnings", "retained_earnings", False),
        ("Total Equity", "total_equity", True),
        ("", None, False),
        ("TOTAL LIABILITIES + EQUITY", "total_liabilities_and_equity", True),
    ]
    r += 1
    for i, (label, attr, is_bold) in enumerate(lines):
        if not attr:
            r += 1
            continue
        vals = [label] + [round(ye.get(y, {}).get(attr, 0), 0) for y in sorted(ye)]
        _row(ws, r, 2, vals, bold=is_bold, nf=FMT_NUM, alt=i % 2 == 0)
        r += 1


def _build_cash_flow(ws, proj, cfg):
    ws.title = "Cash_Flow"
    ws.sheet_properties.tabColor = "BF8F00"
    _col(ws, {1: 3, 2: 30, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16})

    ws.merge_cells('B2:G2')
    ws['B2'] = f"{cfg.company} — Cash Flow Statement (5-Year Projection)"
    ws['B2'].font = FMT_TITLE

    heads = ["Metric"] + [str(y) for y in sorted(set(p.year for p in proj.periods))]
    r = 4
    _hdr(ws, r, 2, heads)

    ann_cf = _annualize_cf(proj)

    lines = [
        ("OPERATING ACTIVITIES", None, True),
        ("Net Income", "net_income", False),
        ("Depreciation", "depreciation", False),
        ("Change in AR", "change_in_ar", False),
        ("Change in Inventory", "change_in_inventory", False),
        ("Change in AP", "change_in_ap", False),
        ("Net Operating Cash Flow", "operating_cash_flow", True),
        ("", None, False),
        ("INVESTING ACTIVITIES", None, True),
        ("Capital Expenditure", "capex", False),
        ("Net Investing Cash Flow", "investing_cash_flow", True),
        ("", None, False),
        ("FINANCING ACTIVITIES", None, True),
        ("Drawdowns", "drawdowns", False),
        ("Repayments", "repayments", False),
        ("Dividends", "dividends", False),
        ("Net Financing Cash Flow", "financing_cash_flow", True),
        ("", None, False),
        ("Net Change in Cash", "net_change_in_cash", True),
        ("Opening Cash", "opening_cash", False),
        ("Closing Cash", "closing_cash", True),
    ]
    r += 1
    for i, (label, attr, is_bold) in enumerate(lines):
        if not attr:
            r += 1
            continue
        vals = [label] + [round(ann_cf.get(y, {}).get(attr, 0), 0) for y in sorted(ann_cf)]
        _row(ws, r, 2, vals, bold=is_bold, nf=FMT_NUM, alt=i % 2 == 0)
        r += 1


def _build_overview(ws, proj, cfg, audit):
    ws.title = "Overview"
    ws.sheet_properties.tabColor = "002060"
    _col(ws, {1: 3, 2: 30, 3: 50})

    ws.merge_cells('B2:C2')
    ws['B2'] = f"{cfg.company} — Model Overview"
    ws['B2'].font = FMT_TITLE

    meta = [
        ("Company", cfg.company),
        ("Currency", cfg.currency),
        ("Projection Horizon", f"{proj.horizon_years} years ({proj.n_quarters} quarters)"),
        ("Base Year", str(proj.base_year)),
        ("All Periods Balanced", str(proj.all_balanced())),
        ("", ""),
        ("Revenue Year 1", f"ETB {sum(p.income.revenue for p in proj.periods if p.year == proj.base_year):,.0f}"),
        ("Revenue Year 5", f"ETB {sum(p.income.revenue for p in proj.periods if p.year == proj.base_year + proj.horizon_years - 1):,.0f}"),
        ("CAGR", f"{(sum(p.income.revenue for p in proj.periods if p.year == proj.base_year + proj.horizon_years - 1) / max(sum(p.income.revenue for p in proj.periods if p.year == proj.base_year), 1)) ** (1/max(proj.horizon_years - 1, 1)) - 1:.1%}"),
        ("", ""),
    ]
    if audit:
        meta += [
            ("Run ID", audit.get("run_id", "?")),
            ("Scenario", audit.get("scenario", "base")),
            ("Model Version", audit.get("model_version", "?")),
        ]

    r = 4
    for label, value in meta:
        if not label:
            r += 1
            continue
        ws.cell(r, 2, label).font = FMT_BOLD
        ws.cell(r, 3, value).font = Font(size=10, name="Calibri")
        r += 1


# ─── Aggregation helpers ─────────────────────────────────────────────────


def _annualize(proj, statement="income"):
    ann: dict[int, dict] = {}
    for p in proj.periods:
        y = p.year
        if y not in ann:
            ann[y] = {}
        src = p.income if statement == "income" else p.balance
        for k, v in src.__dict__.items():
            ann[y][k] = ann[y].get(k, 0.0) + v
    return ann


def _year_end(proj):
    """Return Q4 (year-end) balances for each year."""
    ye: dict[int, dict] = {}
    for p in proj.periods:
        if p.quarter == 4:
            ye[p.year] = dict(p.balance.__dict__)
    return ye


def _annualize_cf(proj):
    ann: dict[int, dict] = {}
    for p in proj.periods:
        y = p.year
        if y not in ann:
            ann[y] = {}
        for k, v in p.cash_flow.__dict__.items():
            ann[y][k] = ann[y].get(k, 0.0) + v
    return ann
