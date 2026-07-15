"""Configuration loader — YAML or Excel input template"""
from __future__ import annotations
import yaml
import json
from pathlib import Path
from jsonschema import validate as json_validate, ValidationError
from typing import Any, Optional
from dataclasses import dataclass, field
from datetime import date

from yusra_model.domain import (
    StrategicContext, BusinessTargets,
    RevenueDrivers, ProductLine, SeasonalityProfile, CustomerSegment,
    CostStructure, COGSBreakdown, OperatingExpenses, OpExCategory, HeadcountPlan,
    EquityPosition, DebtFacility,
    DepreciableAsset, CapExPlan, FixedAssets,
    InventoryPolicy, ReceivablesPolicy, PayablesPolicy, WorkingCapitalPolicy,
    TaxSettings,
)


@dataclass
class Config:
    """Normalised configuration object — legacy + expanded business domains.

    All new domain fields default to ``None`` so existing configs load unchanged.
    The legacy flat fields (``opening_cash``, ``total_facility``, etc.) remain
    at the top level for backward compatibility.
    """
    # ── Legacy fields (unchanged) ────────────────────────────────────────
    company: str
    currency: str
    opening_cash: float
    opening_inventory: float
    total_facility: float
    overheads_per_month: float
    profit_rate: float
    loan_tenor_quarters: int
    baseline_rate: float
    stress_rate: float
    sales_cycle_options: list[str]
    cash_ratio: float
    credit_ratio: float
    gross_profit_margin: float
    loans: list[dict]
    ceo_throughput_target: float
    plan_throughput_target: float
    min_sales_buffer: float

    # ── New domain fields (optional → None when absent) ─────────────────
    strategy: Optional[StrategicContext] = None
    revenue: Optional[RevenueDrivers] = None
    costs: Optional[CostStructure] = None
    capital: list[DebtFacility] = field(default_factory=list)
    equity: Optional[EquityPosition] = None
    fixed_assets: Optional[FixedAssets] = None
    working_capital_policy: Optional[WorkingCapitalPolicy] = None
    taxation: Optional[TaxSettings] = None
    fiscal_year_start_month: int = 1


def load_config(path: str | Path) -> Config:
    """Load YAML config, validate against schema, return normalised Config."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {path}")

    with open(path, encoding="utf-8") as f:
        raw = yaml.safe_load(f)

    schema_path = Path(__file__).parent.parent.parent.parent / "config" / "schema.json"
    with open(schema_path, encoding="utf-8") as f:
        schema = json.load(f)

    try:
        json_validate(raw, schema)
    except ValidationError as e:
        raise ValueError(f"Config validation error: {e.message}") from e

    # ── Parse legacy sections ──
    p = raw["parameters"]
    ex = raw["exchanges"]
    s = raw["sales"]
    t = raw["targets"]

    loans = []
    for loan in raw["loans"]:
        l = dict(loan)
        l["start_date"] = date.fromisoformat(l["start_date"])
        loans.append(l)

    # ── Parse new domain sections (optional ──
    strategy = _parse_strategy(raw.get("strategy"))
    revenue = _parse_revenue(raw.get("revenue"))
    costs = _parse_costs(raw.get("costs"))
    debt_facilities = _parse_debt(raw.get("capital"))
    equity = _parse_equity(raw.get("capital"))
    fixed_assets = _parse_fixed_assets(raw.get("fixed_assets"))
    wc_policy = _parse_working_capital(raw.get("working_capital"))
    taxation = _parse_taxation(raw.get("taxation"))

    return Config(
        # Legacy
        company=raw["company"],
        currency=raw["currency"],
        opening_cash=float(p["opening_cash"]),
        opening_inventory=float(p["opening_inventory"]),
        total_facility=float(p["total_facility"]),
        overheads_per_month=float(p["overheads_per_month"]),
        profit_rate=float(p["profit_rate"]),
        loan_tenor_quarters=int(p["loan_tenor_quarters"]),
        baseline_rate=float(ex["baseline"]),
        stress_rate=float(ex["stress"]),
        sales_cycle_options=list(s["cycle_options"]),
        cash_ratio=float(s["cash_ratio"]),
        credit_ratio=float(s["credit_ratio"]),
        gross_profit_margin=float(s["gross_profit_margin"]),
        loans=loans,
        ceo_throughput_target=float(t.get("ceo_throughput", 240000000)),
        plan_throughput_target=float(t.get("plan_throughput", 140000000)),
        min_sales_buffer=float(t.get("min_sales_buffer", 500000)),
        # New domains
        strategy=strategy,
        revenue=revenue,
        costs=costs,
        capital=debt_facilities,
        equity=equity,
        fixed_assets=fixed_assets,
        working_capital_policy=wc_policy,
        taxation=taxation,
        fiscal_year_start_month=int(p.get("fiscal_year_start_month", 1)),
    )


# ─── Private parsing helpers ───────────────────────────────────────────────


def _parse_strategy(raw: dict | None) -> StrategicContext | None:
    if raw is None:
        return None
    bt = raw.get("targets", {})
    targets = BusinessTargets(
        revenue_growth_annual=float(bt.get("revenue_growth_annual", 0.15)),
        gross_margin=float(bt.get("gross_margin", 0.28)),
        operating_margin=float(bt.get("operating_margin", 0.12)),
        net_margin=float(bt.get("net_margin", 0.08)),
        roe=float(bt.get("roe", 0.25)),
        roic=float(bt.get("roic", 0.20)),
        dscr=float(bt.get("dscr", 1.5)),
        current_ratio=float(bt.get("current_ratio", 2.0)),
        lt_debt_to_ebitda=float(bt.get("lt_debt_to_ebitda", 2.0)),
        debt_to_equity=float(bt.get("debt_to_equity", 1.5)),
        dividend_payout_ratio=float(bt.get("dividend_payout_ratio", 0.30)),
        cash_conversion_cycle_days=int(bt.get("cash_conversion_cycle_days", 60)),
        inventory_turns=float(bt.get("inventory_turns", 6.0)),
    )
    return StrategicContext(
        planning_horizon_years=int(raw.get("planning_horizon_years", 5)),
        base_year=int(raw.get("base_year", 2025)),
        targets=targets,
    )


def _parse_revenue(raw: dict | None) -> RevenueDrivers | None:
    if raw is None:
        return None
    pl_raw = raw.get("product_lines", [])
    product_lines = [
        ProductLine(
            name=pl["name"],
            base_volume=float(pl.get("base_volume", 0)),
            avg_price=float(pl.get("avg_price", 0)),
            growth_rate=float(pl.get("growth_rate", 0)),
            cogs_per_unit=float(pl.get("cogs_per_unit", 0)),
            gross_margin=float(pl.get("gross_margin", 0)),
            description=pl.get("description", ""),
        )
        for pl in pl_raw
    ]
    seas = raw.get("seasonality", {})
    seasonality = SeasonalityProfile(
        q1=float(seas.get("q1", 0.25)),
        q2=float(seas.get("q2", 0.25)),
        q3=float(seas.get("q3", 0.25)),
        q4=float(seas.get("q4", 0.25)),
    )
    segs = raw.get("customer_segments", [])
    segments = [
        CustomerSegment(
            name=cs["name"],
            revenue_pct=float(cs.get("revenue_pct", 0)),
            payment_terms_days=int(cs.get("payment_terms_days", 30)),
            bad_debt_rate=float(cs.get("bad_debt_rate", 0.005)),
        )
        for cs in segs
    ]
    return RevenueDrivers(
        product_lines=product_lines,
        seasonality=seasonality,
        cash_ratio=float(raw.get("cash_ratio", 0.85)),
        credit_ratio=float(raw.get("credit_ratio", 0.15)),
        credit_terms_days=int(raw.get("credit_terms_days", 30)),
        bad_debt_rate=float(raw.get("bad_debt_rate", 0.005)),
        customer_segments=segments,
    )


def _parse_costs(raw: dict | None) -> CostStructure | None:
    if raw is None:
        return None
    cogs = raw.get("cogs_breakdown", {})
    opex = raw.get("operating_expenses", {})
    hc = raw.get("headcount", {})
    return CostStructure(
        cogs_breakdown=COGSBreakdown(
            raw_materials_pct=float(cogs.get("raw_materials_pct", 0.55)),
            import_duties_pct=float(cogs.get("import_duties_pct", 0.15)),
            freight_pct=float(cogs.get("freight_pct", 0.10)),
            direct_labor_pct=float(cogs.get("direct_labor_pct", 0.10)),
            other_direct_pct=float(cogs.get("other_direct_pct", 0.10)),
        ),
        operating_expenses=OperatingExpenses(
            sales_marketing=_parse_opex(opex.get("sales_marketing")),
            distribution=_parse_opex(opex.get("distribution")),
            admin=_parse_opex(opex.get("admin")),
            r_and_d=_parse_opex(opex.get("r_and_d")),
            other=_parse_opex(opex.get("other")),
        ),
        headcount=HeadcountPlan(
            total_headcount=int(hc.get("total_headcount", 0)),
            avg_cost_per_employee_per_month=float(hc.get("avg_cost_per_employee_per_month", 0)),
            hiring_growth_rate=float(hc.get("hiring_growth_rate", 0)),
            salary_escalation_rate=float(hc.get("salary_escalation_rate", 0)),
        ),
        escalation_rate=float(raw.get("escalation_rate", 0.08)),
        other_fixed_costs_per_month=float(raw.get("other_fixed_costs_per_month", 0)),
    )


def _parse_opex(raw: dict | None) -> OpExCategory:
    if raw is None:
        return OpExCategory()
    return OpExCategory(
        fixed_per_month=float(raw.get("fixed_per_month", 0)),
        variable_pct_of_revenue=float(raw.get("variable_pct_of_revenue", 0)),
        description=raw.get("description", ""),
    )


def _parse_debt(raw: dict | None) -> list[DebtFacility]:
    if raw is None:
        return []
    facilities = []
    for d in raw.get("debt", []):
        facilities.append(DebtFacility(
            facility=float(d.get("facility", 0)),
            type=d.get("type", "Murabaha_Revolving"),
            profit_rate=float(d.get("profit_rate", 0.07)),
            tenor_quarters=int(d.get("tenor_quarters", 8)),
            purpose=d.get("purpose", ""),
            outstanding=float(d.get("outstanding", 0)),
        ))
    return facilities


def _parse_equity(raw: dict | None) -> EquityPosition | None:
    if raw is None:
        return None
    eq = raw.get("equity")
    if eq is None:
        return None
    return EquityPosition(
        paid_in_capital=float(eq.get("paid_in_capital", 0)),
        retained_earnings=float(eq.get("retained_earnings", 0)),
        additional_paid_in=float(eq.get("additional_paid_in", 0)),
    )


def _parse_fixed_assets(raw: dict | None) -> FixedAssets | None:
    if raw is None:
        return None
    assets = []
    for a in raw.get("existing_assets", []):
        assets.append(DepreciableAsset(
            cost=float(a.get("cost", 0)),
            accumulated_depreciation=float(a.get("accumulated_depreciation", 0)),
            useful_life_years=int(a.get("useful_life_years", 10)),
            depreciation_method=a.get("depreciation_method", "straight_line"),
            description=a.get("description", ""),
        ))
    cp = raw.get("capex_plan", {})
    capex = CapExPlan(
        year_1=float(cp.get("year_1", 0)),
        year_2=float(cp.get("year_2", 0)),
        year_3=float(cp.get("year_3", 0)),
        year_4=float(cp.get("year_4", 0)),
        year_5=float(cp.get("year_5", 0)),
    )
    return FixedAssets(
        existing_assets=assets,
        capex_plan=capex,
        default_capex_useful_life_years=raw.get("default_capex_useful_life_years", 10),
    )


def _parse_working_capital(raw: dict | None) -> WorkingCapitalPolicy | None:
    if raw is None:
        return None
    inv = raw.get("inventory", {})
    rec = raw.get("receivables", {})
    pay = raw.get("payables", {})
    return WorkingCapitalPolicy(
        inventory=InventoryPolicy(
            raw_materials_days=int(inv.get("raw_materials_days", 45)),
            wip_days=int(inv.get("wip_days", 15)),
            finished_goods_days=int(inv.get("finished_goods_days", 60)),
            safety_stock_days=int(inv.get("safety_stock_days", 15)),
            obsolescence_rate=float(inv.get("obsolescence_rate", 0.02)),
        ),
        receivables=ReceivablesPolicy(
            dso_target=int(rec.get("dso_target", 30)),
            bad_debt_rate=float(rec.get("bad_debt_rate", 0.005)),
        ),
        payables=PayablesPolicy(
            dpo_target=int(pay.get("dpo_target", 45)),
        ),
        cash_conversion_cycle_target=int(raw.get("cash_conversion_cycle_target", 60)),
    )


def _parse_taxation(raw: dict | None) -> TaxSettings | None:
    if raw is None:
        return None
    return TaxSettings(
        corporate_income_tax_rate=float(raw.get("corporate_income_tax_rate", 0.30)),
        vat_rate=float(raw.get("vat_rate", 0.15)),
        withholding_tax_rate=float(raw.get("withholding_tax_rate", 0.05)),
        tax_loss_carryforward_years=int(raw.get("tax_loss_carryforward_years", 5)),
        tax_installment_frequency_months=int(raw.get("tax_installment_frequency_months", 3)),
    )


def generate_input_template(path: str | Path) -> None:
    """Generate a blank Excel input template for non-technical users."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inputs"

    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    hdr_fill = PatternFill("solid", fgColor="003366")
    lbl_font = Font(bold=True, size=10, name="Calibri")
    thin = Border(left=Side("thin", "C0C0C0"), right=Side("thin", "C0C0C0"),
                  top=Side("thin", "C0C0C0"), bottom=Side("thin", "C0C0C0"))

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22

    def hdr(row, cols):
        for i, c in enumerate(cols):
            cell = ws.cell(row, 2 + i, c)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal="center"); cell.border = thin

    r = 1
    ws.cell(r, 2, "YUSRA PHARMA PLC — INPUT TEMPLATE").font = Font(bold=True, size=14, color="003366")

    r = 3
    ws.cell(r, 2, "PARAMETERS").font = Font(bold=True, size=12, color="003366")
    r = 4
    hdr(r, ["Parameter", "Value", "Notes"])
    params = [
        ("Opening Cash (ETB)", 16927876.79, "Current cash balance"),
        ("Opening Inventory (ETB)", 42000000, "Stock on hand"),
        ("Total Facility (ETB)", 70000000, "Murabaha revolving loan"),
        ("Overheads per Month (ETB)", 1444225.35, "Fixed operating costs"),
        ("Profit Rate (annual)", 0.07, "Murabaha flat rate"),
        ("Loan Tenor (quarters)", 8, "Repayment period"),
    ]
    for i, (lbl, val, note) in enumerate(params):
        rr = r + 1 + i
        ws.cell(rr, 2, lbl).font = lbl_font; ws.cell(rr, 2).border = thin
        ws.cell(rr, 3, val).font = lbl_font; ws.cell(rr, 3).border = thin
        ws.cell(rr, 4, note).font = Font(size=9, italic=True, color="888888")

    r += len(params) + 2
    ws.cell(r, 2, "EXCHANGE RATES").font = Font(bold=True, size=12, color="003366")
    r += 1
    hdr(r, ["Scenario", "Rate", "Active?", "Notes"])
    r += 1
    ws.cell(r, 2, "Baseline").font = lbl_font; ws.cell(r, 3, 160); ws.cell(r, 4, "Yes"); ws.cell(r, 5, "Current rate")
    r += 1
    ws.cell(r, 2, "Stress").font = lbl_font; ws.cell(r, 3, 175); ws.cell(r, 4, ""); ws.cell(r, 5, "Depreciation scenario")

    r += 2
    ws.cell(r, 2, "SALES ASSUMPTIONS").font = Font(bold=True, size=12, color="003366")
    r += 1
    hdr(r, ["Parameter", "Value", "Notes"])
    sales_items = [
        ("Sales Cycle", "3 Months", "3 Months or 6 Months"),
        ("Cash Sales %", 0.85, "Immediate collection"),
        ("Credit Sales %", 0.15, "30-day lag"),
        ("Gross Profit Margin", 0.30, "Markup on COGS"),
    ]
    for i, (lbl, val, note) in enumerate(sales_items):
        rr = r + 1 + i
        ws.cell(rr, 2, lbl).font = lbl_font; ws.cell(rr, 2).border = thin
        ws.cell(rr, 3, val).font = lbl_font; ws.cell(rr, 3).border = thin
        ws.cell(rr, 4, note).font = Font(size=9, italic=True, color="888888")

    r += len(sales_items) + 2
    ws.cell(r, 2, "LOAN PIPELINE").font = Font(bold=True, size=12, color="003366")
    r += 1
    hdr(r, ["Supplier", "USD Value", "ETB Principal", "Start Date", "Quarterly Repayment"])
    sample_loans = [
        ("Reyoung (CHINA)", 147354, 23580195.28, "2026-04-07", 3360178),
        ("SCOTT EDIL (INDIA)", 194100, "", "2026-10-01", ""),
        ("TSM (MALAYSIA)", 42840, "", "2026-10-10", ""),
        ("Tinachin HENGSHENG", 61904, "", "2026-10-10", ""),
    ]
    for i, (sup, usd, etb, sdt, qpmt) in enumerate(sample_loans):
        rr = r + 1 + i
        ws.cell(rr, 2, sup).font = lbl_font; ws.cell(rr, 2).border = thin
        ws.cell(rr, 3, usd).border = thin
        ws.cell(rr, 4, etb).border = thin
        ws.cell(rr, 5, sdt).border = thin
        ws.cell(rr, 6, qpmt).border = thin

    r += len(sample_loans) + 2
    ws.cell(r, 2, "TARGETS").font = Font(bold=True, size=12, color="003366")
    r += 1
    hdr(r, ["Target", "Value (ETB)", "Notes"])
    targets = [
        ("CEO Throughput Goal", 240000000, "3.4x facility utilisation"),
        ("Plan Throughput", 140000000, "2.0x facility utilisation"),
    ]
    for i, (lbl, val, note) in enumerate(targets):
        rr = r + 1 + i
        ws.cell(rr, 2, lbl).font = lbl_font; ws.cell(rr, 2).border = thin
        ws.cell(rr, 3, val).font = lbl_font; ws.cell(rr, 3).border = thin
        ws.cell(rr, 4, note).font = Font(size=9, italic=True, color="888888")

    wb.save(path)
    print(f"Template saved: {path}")
